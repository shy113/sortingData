# *-* encoding:utf8 *_*
# Author       : GZH
# Date         : 2022年11月02日
# Description  : 升级版
import os

import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook

from sortingData.classes.table import Product, Feature
from sortingData.classes.merge import Merage


# 拿到指定表
def getTable(filename, sheetname):
    xlsx = xlrd.open_workbook(filename)
    # print("sheets：" + str(xlsx.sheet_names()))
    table_data = xlsx.sheet_by_name(sheetname)
    return table_data


# 获取表中所有数据 封装成对象形式 塞在列表里面
def getAllDatalist(table_data):
    '''
    构造表中所有数据 列表
    :return: 构造好的列表
    '''
    list = []
    effectiveRow, effectiveCol = table_data.nrows, table_data.ncols
    for ri in range(effectiveRow):
        rowDatalist = []
        for ci in range(effectiveCol):
            # 产品属性
            rowDatalist.append(table_data.cell_value(ri, ci))
        product = Product(tuple(rowDatalist))
        list.append(product)
    return list


# 获取特征列表
def getFeatureDicList(datalist):
    featureList = []
    for data in datalist:
        data: Product = data
        feature_tuple = (data.auxiliary_attribute_category_FName, data.default_repository_FName)
        feature = Feature(feature_tuple)
        feature_dict = feature.getObjToDict()
        # 对象比较是比较他们的地址 字典比较是比较他们的内容
        if feature_dict not in featureList:
            featureList.append(feature_dict)
    # print(len(featureList))
    return featureList


# 重新排列表中数据
def updateList(datalist, featurelist):
    for data in datalist:
        data: Product = data  # 并不是新建了一个对象 只是引用原对象的地址
        data.featureId = featurelist.index(data.getFeatureDict())  # 通过内容拿到索引
    # 此刻 datalist 中的每一个对象的 featureId 都有了值
    Merage().merge_sort(datalist)  # 调用归并排序 (稳定性高)
    for data in datalist:
        print(data.featureId)
    # print(datalist)
    return datalist


# 将修改好的数据写入表中
def write_excel_xls(list, filename, sheetname):
    '''
    将列表数据写入excel表格
    :param list: 存放数据的列表
    :param sheetname: 表名
    :param filename: 文件名
    :return:
    '''
    if not os.path.exists(filename):
        wb = Workbook(filename)
        wb.save(filename)
    wb = load_workbook(filename)
    if sheetname not in wb.sheetnames:
        wb.create_sheet(sheetname)
    ws = wb[sheetname]
    r = 1  # 起始单元格是(1,1)而不是(0,0)
    for item in list:
        c = 1
        item: Product = item
        dict = item.getObjToDict()
        for k, v in dict.items():
            ws.cell(r, c, v)
            c += 1
        r += 1
    wb.save(filename)


if __name__ == '__main__':
    dir = os.getcwd()  # 当前目录
    filename, sheetname, newsheetname = dir + './开发部__面试__题目.xlsx', '分类数据', 'newData'
    table_data = getTable(filename, sheetname)
    datalist = getAllDatalist(table_data)
    featurediclist = getFeatureDicList(datalist)
    newData_list = updateList(datalist, featurediclist)
    write_excel_xls(newData_list, filename, newsheetname)
